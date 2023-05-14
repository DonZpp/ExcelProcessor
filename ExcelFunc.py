import os
import time
import openpyxl


# 课程表
strScheduleDir = 'D:\document\GanDong Acadamy\中职\课表'
strScheduleName = '2023下半学期单双周课表.xlsx'
strScheduleSheetName = '4.13春季课表'

# 教师表
strTchWBDir = 'D:\document\GanDong Acadamy\中职'
strTchWBName = '教师名单.xlsx'
strTchSheetName = 'Sheet1'

# 输入表格
strInitDir = 'D:\document-unimportant\ExcelWorkBench'
strInitTable = '初始数据.xlsx'
strInitDataSheetName = 'Sheet'

#单元格格式数据
sInitDataFont = openpyxl.styles.Font(size=20)
nClsNameCellWidth = 40


# test = openpyxl.Workbook()
# for sheet in test:
#     print(sheet.title)
# sheet = test['Sheet']
# sheet[1][0].value = 'hello'


nMinRow = 4
nMaxRow = 24
nMinCol = 1
nMaxCol = 40
strEmpty = 'Empty'


def OpenScheduleWB():
    os.chdir(strScheduleDir)
    return openpyxl.load_workbook(strScheduleName)


def OpenTchWB():
    os.chdir(strTchWBDir)
    return openpyxl.load_workbook(strTchWBName)


def ReadTeacherName():
    workbook = OpenTchWB()
    arrTeacherName = {}
    sTeachSheet = workbook[strTchSheetName]
    i = 0
    for sCols in sTeachSheet[sTeachSheet.dimensions]:
        for cell in sCols:
            arrTeacherName[i] = cell.value
            i = i + 1

    return arrTeacherName


arrTchs = ReadTeacherName()


def ColIndex2Num(strIdx):
    # column 1 -> A, 2 -> B
    return openpyxl.utils.column_index_from_string(strIdx)


def GetCurrKey(strCurrName, strClsName):
    return strCurrName + strClsName


class CurrInfo:
    def __init__(self, strCurrName, strClsName, strTchName) -> None:
        self.strCurrName = strCurrName
        self.strClsName = strClsName
        self.strTchName = strTchName
        self.nCount = 1

    def IsSame(self, cr):
        return self.strCurrName == cr.strCurrName and self.strClsName == cr.strClsName

    def GetKey(self):
        return GetCurrKey(self.strCurrName, self.strClsName)


class TchData:
    def __init__(self, strTchName, strClsName, strCurrName):
        self.strTchName = strTchName
        self.dicCurrs = {}
        self.dicCurrs[GetCurrKey(strCurrName, strClsName)] = CurrInfo(strCurrName, strClsName, strTchName)


class ClsData:
    def __init__(self, strTchName, strClsName, strCurrName):
        self.strClsName = strClsName
        self.dicCurrs = {}
        self.dicCurrs[GetCurrKey(strCurrName, strClsName)] = CurrInfo(strCurrName, strClsName, strTchName)


def Space_NextlineFilter(strOrg):
    if (type(strOrg) == type('')):
        strTmp = strOrg.replace(' ', '')
        strTmp = strTmp.replace('\n', '')
        strTmp = strTmp.replace('\r', '')
        return strTmp
    else:
        return strOrg


def CellVal(cell):
    return Space_NextlineFilter(cell.value)


# ret CurrName, TchName
def GetCurrNameAndTchName(strOrgName):
    if type(strOrgName) !=  type(''):
        return strEmpty, strEmpty
    i = 0
    nTchNum = len(arrTchs)
    while i < len(arrTchs):
        strTchName = arrTchs[i]
        nPos = strOrgName.find(strTchName)
        if nPos != -1:
            return strOrgName[0: nPos], strTchName
        i = i + 1
    
    return strOrgName, strEmpty


def IncCurr(TchMapCurr, ClsMapCurr, strCurrName, strTchName, strClsName):
    #print(strTchName, strClsName, strCurrName)
    # create if not exist
    # else increase count
    if not (strTchName in TchMapCurr):
        TchMapCurr[strTchName] = TchData(strTchName, strClsName, strCurrName)
    elif not (GetCurrKey(strCurrName, strClsName) in TchMapCurr[strTchName].dicCurrs):
        TchMapCurr[strTchName].dicCurrs[GetCurrKey(strCurrName, strClsName)] = CurrInfo(strCurrName, strClsName, strTchName)
    else:
        sCurrInfo = TchMapCurr[strTchName].dicCurrs[GetCurrKey(strCurrName, strClsName)]
        sCurrInfo.nCount = sCurrInfo.nCount + 1

    if not (strClsName in ClsMapCurr):
        ClsMapCurr[strClsName] = ClsData(strTchName, strClsName, strCurrName)
    elif not (GetCurrKey(strCurrName, strClsName) in ClsMapCurr[strClsName].dicCurrs):
        ClsMapCurr[strClsName].dicCurrs[GetCurrKey(strCurrName, strClsName)] = CurrInfo(strCurrName, strClsName, strTchName)
    else:
        sCurrInfo = ClsMapCurr[strClsName].dicCurrs[GetCurrKey(strCurrName, strClsName)]
        sCurrInfo.nCount = sCurrInfo.nCount + 1

    return


def PrintCurrData(CurrMap):
    for sCurrInfos in CurrMap.values():
        for sCI in sCurrInfos.dicCurrs.values():
            if (sCI.strTchName == strEmpty and (sCI.strCurrName != '自习' and sCI.strCurrName != '班会')):
                print('error!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
            print(sCI.strClsName, sCI.strTchName, sCI.strCurrName, sCI.nCount)


def WriteInitData(ClsMapCurr):
    os.chdir(strInitDir)
    sInitBook = openpyxl.Workbook()
    sheet = sInitBook[strInitDataSheetName]

    # column width
    sheet.column_dimensions['A'].width = nClsNameCellWidth
    sheet.column_dimensions['B'].width = nClsNameCellWidth
    sheet.column_dimensions['C'].width = nClsNameCellWidth

    i = 1
    for sCurrData in ClsMapCurr.values():
        for sCI in  sCurrData.dicCurrs.values():
            # cell value
            sheet.cell(i, 1).value = sCI.strClsName
            sheet.cell(i, 2).value = sCI.strTchName
            sheet.cell(i, 3).value = sCI.strCurrName
            sheet.cell(i, 4).value = sCI.nCount
            # font size
            sheet.cell(i, 1).font = sInitDataFont
            sheet.cell(i, 2).font = sInitDataFont
            sheet.cell(i, 3).font = sInitDataFont
            sheet.cell(i, 4).font = sInitDataFont
            i = i + 1
    sInitBook.save(strInitTable)


def ScheduleStatistic():
    workbook = OpenScheduleWB()

    TchMapCurr = {}     # 教师及其课程
    ClsMapCurr = {}     # 班级及其课程

    # 遍历课程表
    sheet = workbook[strScheduleSheetName]
    i = nMinRow
    j = nMinCol
    while i <= nMaxRow:
        while j <= nMaxCol:
            # inc class count
            strClsName = CellVal(sheet[i][0])
            strCurrName, strTchName = GetCurrNameAndTchName(CellVal(sheet[i][j]))
            if strCurrName != strEmpty:
                IncCurr(TchMapCurr, ClsMapCurr, strCurrName, strTchName, strClsName)
            j = j + 1
        i = i + 1
        j = nMinCol





#ReadTeacherName()
ScheduleStatistic()