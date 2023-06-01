import os
import openpyxl
import Common


def AddCurr(TchMapCurr, ClsMapCurr, strClsName, strTchName, strCurrName, nCount, strIsConsecutive, strBanList):
    sCI = Common.InitCurrInfo(strCurrName, strClsName, strTchName, nCount, strIsConsecutive, strBanList)

    if not (strTchName in TchMapCurr):
        TchMapCurr[strTchName] = Common.TchData(strTchName, sCI)
    elif not (sCI.GetKey() in TchMapCurr[strTchName].dicCurrs):
        TchMapCurr[strTchName].dicCurrs[sCI.GetKey()] = sCI 

    if (not strClsName in ClsMapCurr): 
        ClsMapCurr[strClsName] = Common.ClsData(strClsName, sCI)
    elif not (sCI.GetKey() in ClsMapCurr[strClsName].dicCurrs):
        ClsMapCurr[strClsName].dicCurrs[sCI.GetKey()] = sCI 



def CreateCurrOrder(ClsMapCurr):
    CurrInfoArr = {}
    i = 1
    for sClsData in ClsMapCurr.values():
        for sCI in sClsData.dicCurrs.values():
            CurrInfoArr[i] = sCI
            i = i + 1
    return CurrInfoArr


def ReadInitTable():
    os.chdir(Common.strInitDir)
    wb = openpyxl.load_workbook(Common.strInitTable)
    sheet = wb[Common.strInitDataSheetName]
    i = Common.nInitDataRowBeg

    ClsMapCurr = {}
    TchMapCurr = {}

    while (not sheet.cell(i, Common.INIT_CLS_INDEX).value == None \
            and not sheet.cell(i, Common.INIT_TCH_INDEX).value == None \
            and not sheet.cell(i, Common.INIT_CURR_INDEX).value == None):
        strClsName = sheet.cell(i, Common.INIT_CLS_INDEX).value
        strTchName = sheet.cell(i, Common.INIT_TCH_INDEX).value
        strCurrName = sheet.cell(i, Common.INIT_CURR_INDEX).value
        nCount = sheet.cell(i, Common.INIT_COUNT_INDEX).value
        strIsConsecutive = sheet.cell(i, Common.INIT_CONSECUTIVE_INDEX).value
        strBanList = sheet.cell(i, Common.INIT_BAN_INDEX).value
        AddCurr(TchMapCurr, ClsMapCurr, strClsName, strTchName, strCurrName, nCount, strIsConsecutive, strBanList)
        i = i + 1

    CurrInfoArr = CreateCurrOrder(ClsMapCurr)

    return ClsMapCurr, TchMapCurr, CurrInfoArr


# @input CurrInfoArr: ordered curriculums to iterate
# @input nPreCurrIndex: the index of CurrInfoArr point out the curriculum which one was just arranged
# @input ClsMapCurr: for update ban list
# @input TchMapCurr: for update ban list
# @output sSchedule: schedule where curriculum filled in.
def Arrange(CurrInfoArr, nPreCurrIndex, ClsMapCurr, TchMapCurr, sSchedule):
    pass


def GenerateSchedule():
    ClsMapCurr, TchMapCurr, CurrInfoArr = ReadInitTable()


GenerateSchedule()







