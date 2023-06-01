import os
import openpyxl
import Common


def OpenScheduleWB():
    os.chdir(Common.strScheduleDir)
    return openpyxl.load_workbook(Common.strScheduleName)


# ret CurrName, TchName
def GetCurrNameAndTchName(strOrgName):
    if type(strOrgName) !=  type(''):
        return Common.strEmpty, Common.strEmpty
    i = 0
    nTchNum = len(Common.arrTchs)
    while i < len(Common.arrTchs):
        strTchName = Common.arrTchs[i]
        nPos = strOrgName.find(strTchName)
        if nPos != -1:
            return strOrgName[0: nPos], strTchName
        i = i + 1
    
    return strOrgName, Common.strEmpty


def IncCurr(TchMapCurr, ClsMapCurr, strCurrName, strTchName, strClsName):
    # create if not exist, else increase count
    sCI = None
    if not (strTchName in TchMapCurr):
        sCI = Common.CurrInfo(strCurrName, strClsName, strTchName)
        TchMapCurr[strTchName] = Common.TchData(strTchName, sCI)
        if (not strClsName in ClsMapCurr): 
            ClsMapCurr[strClsName] = Common.ClsData(strClsName, sCI)
        else:
            ClsMapCurr[strClsName].dicCurrs[sCI.GetKey()] = sCI 
        return sCI
    elif not (Common.GetCurrKey(strCurrName, strClsName) in TchMapCurr[strTchName].dicCurrs):
        sCI = Common.CurrInfo(strCurrName, strClsName, strTchName)
        TchMapCurr[strTchName].dicCurrs[sCI.GetKey()] = sCI 
        if (not strClsName in ClsMapCurr): 
            ClsMapCurr[strClsName] = Common.ClsData(strClsName, sCI)
        else:
            ClsMapCurr[strClsName].dicCurrs[sCI.GetKey()] = sCI 
        return sCI
    else:
        sCI = TchMapCurr[strTchName].dicCurrs[Common.GetCurrKey(strCurrName, strClsName)]
        sCI.nCount = sCI.nCount + 1
        return None


def WriteInitData(ClsMapCurr):
    os.chdir(Common.strInitDir)
    sInitBook = openpyxl.load_workbook(Common.strInitTable)
    if not Common.strStatisticSheet in sInitBook.sheetnames:
        sInitBook.create_sheet(Common.strStatisticSheet)
    sheet = sInitBook[Common.strStatisticSheet]

    # column width
    sheet.column_dimensions['A'].width = Common.nClsNameCellWidth
    sheet.column_dimensions['B'].width = Common.nClsNameCellWidth
    sheet.column_dimensions['C'].width = Common.nClsNameCellWidth

    i = 2
    for sCurrData in ClsMapCurr.values():
        for sCI in  sCurrData.dicCurrs.values():
            # cell value
            sheet.cell(i, Common.INIT_CLS_INDEX).value = sCI.strClsName
            sheet.cell(i, Common.INIT_TCH_INDEX).value = sCI.strTchName
            sheet.cell(i, Common.INIT_CURR_INDEX).value = sCI.strCurrName
            sheet.cell(i, Common.INIT_COUNT_INDEX).value = sCI.nCount
            # font size
            sheet.cell(i, Common.INIT_CLS_INDEX).font = Common.sInitDataFont
            sheet.cell(i, Common.INIT_TCH_INDEX).font = Common.sInitDataFont
            sheet.cell(i, Common.INIT_CURR_INDEX).font = Common.sInitDataFont
            sheet.cell(i, Common.INIT_COUNT_INDEX).font = Common.sInitDataFont
            i = i + 1
    sInitBook.save(Common.strInitTable)


def ScheduleStatistic():
    workbook = OpenScheduleWB()

    TchMapCurr = {}     # 教师及其课程
    ClsMapCurr = {}     # 班级及其课程

    # 遍历课程表
    sheet = workbook[Common.strScheduleSheetName]
    i = Common.nOldTableMinRow
    j = Common.nOldTableMinCol
    while i <= Common.nOldTableMaxRow:
        while j <= Common.nOldTableMaxCol:
            # inc class count
            strClsName = Common.CellVal(sheet[i][0])
            strCurrName, strTchName = GetCurrNameAndTchName(Common.CellVal(sheet[i][j]))
            if strCurrName != Common.strEmpty:
                IncCurr(TchMapCurr, ClsMapCurr, strCurrName, strTchName, strClsName)
            j = j + 1
        i = i + 1
        j = Common.nOldTableMinCol

    WriteInitData(ClsMapCurr)







